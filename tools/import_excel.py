#!/usr/bin/env python3
"""Convierte fichas Excel (.xlsm) de Anima Beyond Fantasy a actores JSON de animabf."""
from __future__ import annotations

import argparse
import copy
import json
import re
import sys
from pathlib import Path

import openpyxl

MODULE_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = MODULE_ROOT / "data/reference/animabf-template.json"

# Cell map for the "Resumen" sheet (standard Anima Excel character sheet)
CELLS = {
    "name": "M3",
    "level": "F11",
    "class": "L11",
    "life_points": "E12",
    "category": "L12",
    "strength": "F13",
    "dexterity": "J13",
    "agility": "N13",
    "constitution": "R13",
    "power": "F14",
    "intelligence": "J14",
    "willpower": "N14",
    "perception": "R14",
    "presence": "H15",
    "rf": "F16",
    "rm": "J16",
    "rp": "N16",
    "rv": "R16",
    "re": "V16",
    "initiative": "H19",
    "attack": "H22",
    "defense": "H25",
    "damage": "H28",
    "ta_label": "H31",
    "ta_fil": "H32",
    "ta_con": "L32",
    "ta_pen": "P32",
    "ta_cal": "T32",
    "ta_ele": "H33",
    "ta_fri": "L33",
    "ta_ene": "P33",
    "ki_points": "I35",
    "ki_fue": "M37",
    "ki_des": "Q37",
    "ki_agi": "U37",
    "ki_con": "M38",
    "ki_pod": "Q38",
    "ki_vol": "U38",
    "ki_skills": "H39",
    "techniques": "H43",
    "magic_projection": "J48",
    "zeon": "V48",
    "act": "F50",
    "summon": "S50",
    "control": "X50",
    "bind": "S51",
    "banish": "X51",
    "magic_level": "G52",
    "cv": "H60",
    "innate": "O60",
    "psychic_potential": "J62",
    "psychic_projection": "K64",
    "psychic_disciplines": "H66",
    "psychic_powers": "H69",
    "advantages": "K73",
    "natural_abilities": "K76",
    "special": "G79",
    "essential_abilities": "K84",
    "powers": "K87",
    "size": "G93",
    "movement": "S93",
    "regeneration": "Q95",
    "fatigue": "H95",
    "secondary_skills": "D98",
}


def safe_int(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    s = re.sub(r"\s+", "", s)
    m = re.match(r"-?\d+", s)
    return int(m.group(0)) if m else 0


def safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_first_int(value) -> int:
    s = safe_str(value)
    s = re.sub(r"\s+", "", s)
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else 0


def cell(ws, ref):
    return ws[ref].value


def calculate_attribute_modifier(value: int) -> int:
    if value <= 0:
        return 0
    if value < 4:
        return value * 10 - 40
    return min(
        (((value + 5) // 5) + ((value + 4) // 5) + ((value + 2) // 5) - 4) * 5,
        45,
    )


SECONDARY_SKILL_MAP = {
    "acrobacias": ("secondaries", "athletics", "acrobatics"),
    "atletismo": ("secondaries", "athletics", "athleticism"),
    "montar": ("secondaries", "athletics", "ride"),
    "nadar": ("secondaries", "athletics", "swim"),
    "trepar": ("secondaries", "athletics", "climb"),
    "saltar": ("secondaries", "athletics", "jump"),
    "pilotar": ("secondaries", "athletics", "piloting"),
    "frialdad": ("secondaries", "vigor", "composure"),
    "proezas de fuerza": ("secondaries", "vigor", "featsOfStrength"),
    "res. dolor": ("secondaries", "vigor", "withstandPain"),
    "resistir el dolor": ("secondaries", "vigor", "withstandPain"),
    "advertir": ("secondaries", "perception", "notice"),
    "buscar": ("secondaries", "perception", "search"),
    "rastrear": ("secondaries", "perception", "track"),
    "animales": ("secondaries", "intellectual", "animals"),
    "ciencia": ("secondaries", "intellectual", "science"),
    "ley": ("secondaries", "intellectual", "law"),
    "herbolaria": ("secondaries", "intellectual", "herbalLore"),
    "historia": ("secondaries", "intellectual", "history"),
    "tactica": ("secondaries", "intellectual", "tactics"),
    "medicina": ("secondaries", "intellectual", "medicine"),
    "memorizar": ("secondaries", "intellectual", "memorize"),
    "navegacion": ("secondaries", "intellectual", "navigation"),
    "ocultismo": ("secondaries", "intellectual", "occult"),
    "tasacion": ("secondaries", "intellectual", "appraisal"),
    "v. magica": ("secondaries", "intellectual", "magicAppraisal"),
    "valoracion magica": ("secondaries", "intellectual", "magicAppraisal"),
    "estilo": ("secondaries", "social", "style"),
    "intimidar": ("secondaries", "social", "intimidate"),
    "liderazgo": ("secondaries", "social", "leadership"),
    "persuasion": ("secondaries", "social", "persuasion"),
    "comercio": ("secondaries", "social", "trading"),
    "callejeo": ("secondaries", "social", "streetwise"),
    "etiqueta": ("secondaries", "social", "etiquette"),
    "cerrajeria": ("secondaries", "subterfuge", "lockPicking"),
    "disfraz": ("secondaries", "subterfuge", "disguise"),
    "ocultarse": ("secondaries", "subterfuge", "hide"),
    "robo": ("secondaries", "subterfuge", "theft"),
    "sigilo": ("secondaries", "subterfuge", "stealth"),
    "tramperia": ("secondaries", "subterfuge", "trapLore"),
    "venenos": ("secondaries", "subterfuge", "poisons"),
    "arte": ("secondaries", "creative", "art"),
    "baile": ("secondaries", "creative", "dance"),
    "forja": ("secondaries", "creative", "forging"),
    "runas": ("secondaries", "creative", "runes"),
    "alquimia": ("secondaries", "creative", "alchemy"),
    "animismo": ("secondaries", "creative", "animism"),
    "musica": ("secondaries", "creative", "music"),
    "trucos de manos": ("secondaries", "creative", "sleightOfHand"),
    "orfebreria": ("secondaries", "creative", "jewelry"),
    "confeccion": ("secondaries", "creative", "tailoring"),
}


def normalize_skill_name(name):
    import unicodedata
    name = unicodedata.normalize("NFD", name)
    name = "".join(c for c in name if unicodedata.category(c) != "Mn")
    name = name.lower().strip()
    # Remove parenthetical notes like (Ciudad natal)
    name = re.sub(r"\([^)]*\)", "", name).strip()
    return name


def parse_secondary_skills(raw):
    skills = {}
    if not raw:
        return skills
    for part in re.split(r",\s*", str(raw)):
        part = part.strip()
        m = re.match(r"(.+?)\s+(-?\d+)", part)
        if not m:
            continue
        name = normalize_skill_name(m.group(1))
        value = int(m.group(2))
        if name in SECONDARY_SKILL_MAP:
            path = SECONDARY_SKILL_MAP[name]
            skills[path] = value
    return skills


def parse_excel(filepath: Path) -> dict:
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    if "Resumen" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"No 'Resumen' sheet in {filepath.name}")

    ws = wb["Resumen"]
    data = {key: cell(ws, ref) for key, ref in CELLS.items()}
    wb.close()

    return {
        "name": safe_str(data["name"]) or filepath.stem,
        "level": parse_first_int(data["level"]),
        "class": safe_str(data["class"]),
        "category": safe_str(data["category"]),
        "life_points": parse_first_int(data["life_points"]),
        "strength": safe_int(data["strength"]),
        "dexterity": safe_int(data["dexterity"]),
        "agility": safe_int(data["agility"]),
        "constitution": safe_int(data["constitution"]),
        "power": safe_int(data["power"]),
        "intelligence": safe_int(data["intelligence"]),
        "willpower": safe_int(data["willpower"]),
        "perception": safe_int(data["perception"]),
        "presence": safe_int(data["presence"]),
        "rf": safe_int(data["rf"]),
        "rm": safe_int(data["rm"]),
        "rp": safe_int(data["rp"]),
        "rv": safe_int(data["rv"]),
        "re": safe_int(data["re"]),
        "initiative_raw": safe_str(data["initiative"]),
        "initiative": parse_first_int(data["initiative"]),
        "attack_raw": safe_str(data["attack"]),
        "attack": parse_first_int(data["attack"]),
        "defense_raw": safe_str(data["defense"]),
        "defense": parse_first_int(data["defense"]),
        "defense_mode": "block" if "parada" in safe_str(data["defense"]).lower() else "dodge",
        "damage_raw": safe_str(data["damage"]),
        "ta_fil": safe_int(data["ta_fil"]),
        "ta_con": safe_int(data["ta_con"]),
        "ta_pen": safe_int(data["ta_pen"]),
        "ta_cal": safe_int(data["ta_cal"]),
        "ta_ele": safe_int(data["ta_ele"]),
        "ta_fri": safe_int(data["ta_fri"]),
        "ta_ene": safe_int(data["ta_ene"]),
        "zeon": safe_int(data["zeon"]),
        "act": safe_int(data["act"]),
        "magic_projection_raw": safe_str(data["magic_projection"]),
        "magic_projection": parse_first_int(data["magic_projection"]),
        "summon": safe_int(data["summon"]),
        "control": safe_int(data["control"]),
        "bind": safe_int(data["bind"]),
        "banish": safe_int(data["banish"]),
        "cv": safe_int(data["cv"]),
        "innate": safe_int(data["innate"]),
        "psychic_potential": parse_first_int(data["psychic_potential"]),
        "psychic_projection": parse_first_int(data["psychic_projection"]),
        "fatigue": safe_int(data["fatigue"]),
        "size": parse_first_int(data["size"]),
        "movement": parse_first_int(data["movement"]),
        "regeneration": parse_first_int(data["regeneration"]),
        "secondary_skills": parse_secondary_skills(data["secondary_skills"]),
        "advantages": safe_str(data["advantages"]),
        "natural_abilities": safe_str(data["natural_abilities"]),
        "special": safe_str(data["special"]),
        "essential_abilities": safe_str(data["essential_abilities"]),
        "powers": safe_str(data["powers"]),
        "ki_skills": safe_str(data["ki_skills"]),
        "techniques": safe_str(data["techniques"]),
        "magic_level": safe_str(data["magic_level"]),
        "psychic_disciplines": safe_str(data["psychic_disciplines"]),
        "psychic_powers": safe_str(data["psychic_powers"]),
        "source_file": filepath.name,
    }


def build_actor(parsed: dict, template: dict) -> dict:
    from hashlib import sha1

    def stable_id(*parts, length=16):
        return sha1("::".join(parts).encode()).hexdigest()[:length]

    actor_id = stable_id("excel", parsed["source_file"], parsed["name"])
    system = copy.deepcopy(template["Actor"]["character"])

    # Primaries
    for key, val in [
        ("agility", parsed["agility"]),
        ("constitution", parsed["constitution"]),
        ("dexterity", parsed["dexterity"]),
        ("strength", parsed["strength"]),
        ("intelligence", parsed["intelligence"]),
        ("perception", parsed["perception"]),
        ("power", parsed["power"]),
        ("willPower", parsed["willpower"]),
    ]:
        system["characteristics"]["primaries"][key]["value"] = val
        system["characteristics"]["primaries"][key]["mod"] = calculate_attribute_modifier(val)

    level = parsed["level"]
    system["general"]["levels"] = [{
        "_id": stable_id("excel", parsed["name"], "level"),
        "type": "level",
        "name": parsed["name"],
        "flags": {"version": 1},
        "system": {"level": level},
    }]
    system["general"]["presence"]["base"] = {"value": parsed["presence"] or (20 if level <= 0 else 25 + level * 5)}
    system["general"]["aspect"]["race"]["value"] = parsed["class"]
    system["general"]["aspect"]["size"]["value"] = str(parsed["size"])
    system["general"]["description"]["value"] = f"<p>Importado desde {parsed['source_file']}</p>"

    # Secondaries
    lp = parsed["life_points"]
    system["characteristics"]["secondaries"]["lifePoints"]["value"] = lp
    system["characteristics"]["secondaries"]["lifePoints"]["max"] = lp
    system["characteristics"]["secondaries"]["fatigue"]["value"] = parsed["fatigue"]
    system["characteristics"]["secondaries"]["fatigue"]["max"] = parsed["fatigue"]
    system["characteristics"]["secondaries"]["initiative"]["base"]["value"] = parsed["initiative"]
    system["characteristics"]["secondaries"]["movementType"]["mod"]["value"] = parsed["movement"] - parsed["agility"]

    # Resistances
    for res_key, val in [
        ("physical", parsed["rf"]),
        ("magic", parsed["rm"]),
        ("psychic", parsed["rp"]),
        ("disease", parsed["rv"]),
        ("poison", parsed["re"]),
    ]:
        system["characteristics"]["secondaries"]["resistances"][res_key]["base"]["value"] = val

    # Secondary skills
    for path, val in parsed["secondary_skills"].items():
        node = system
        for key in path:
            node = node[key]
        node["base"] = {"value": val}

    # Combat
    system["combat"]["attack"]["base"]["value"] = parsed["attack"]
    if parsed["defense_mode"] == "dodge":
        system["combat"]["dodge"]["base"]["value"] = parsed["defense"]
    else:
        system["combat"]["block"]["base"]["value"] = parsed["defense"]

    # Mystic
    if parsed["zeon"] or parsed["act"] or parsed["magic_projection"]:
        system["ui"]["tabVisibility"]["mystic"]["value"] = True
        system["mystic"]["zeon"]["value"] = parsed["zeon"]
        system["mystic"]["zeon"]["max"] = parsed["zeon"]
        system["mystic"]["act"]["main"]["base"]["value"] = parsed["act"]
        system["mystic"]["magicProjection"]["base"]["value"] = parsed["magic_projection"]
        for key, val in [("summon", parsed["summon"]), ("control", parsed["control"]),
                         ("bind", parsed["bind"]), ("banish", parsed["banish"])]:
            if val:
                system["mystic"]["summoning"][key]["base"]["value"] = val

    # Psychic
    if parsed["psychic_potential"] or parsed["cv"]:
        system["ui"]["tabVisibility"]["psychic"]["value"] = True
        system["psychic"]["psychicPotential"]["base"]["value"] = parsed["psychic_potential"]
        system["psychic"]["psychicProjection"]["base"]["value"] = parsed["psychic_projection"]
        system["psychic"]["psychicPoints"]["value"] = parsed["cv"]
        system["psychic"]["psychicPoints"]["max"] = parsed["cv"]
        system["psychic"]["innatePsychicPower"]["amount"]["value"] = parsed["innate"]

    # Ki / Domine
    if parsed["ki_skills"] or parsed["techniques"]:
        system["ui"]["tabVisibility"]["domine"]["value"] = True

    # Items (armor)
    items = []
    ta_vals = {
        "cut": parsed["ta_fil"], "impact": parsed["ta_con"], "thrust": parsed["ta_pen"],
        "heat": parsed["ta_cal"], "electricity": parsed["ta_ele"],
        "cold": parsed["ta_fri"], "energy": parsed["ta_ene"],
    }
    if any(v > 0 for v in ta_vals.values()):
        armor_id = stable_id(actor_id, "armor", "0")
        armor = {
            "_id": armor_id,
            "name": "Armadura",
            "type": "armor",
            "img": "icons/equipment/chest/breastplate-cuirass-steel-grey.webp",
            "effects": [], "folder": None, "sort": 0, "flags": {},
            "ownership": {"default": 0},
            "_stats": {"systemId": "animabf", "systemVersion": "2.2.1", "coreVersion": "14.359",
                       "createdTime": 1713571200000, "modifiedTime": 1713571200000, "lastModifiedBy": "animuExxetGen001"},
            "_key": f"!actors.items!{actor_id}.{armor_id}",
            "system": {
                **{k: {"base": {"value": v}, "final": {"value": 0}, "value": v} for k, v in ta_vals.items()},
                "pierce": {"base": {"value": 0}, "final": {"value": 0}, "value": 0},
                "integrity": {"base": {"value": 0}, "final": {"value": 0}},
                "presence": {"base": {"value": 0}, "final": {"value": 0}},
                "movementRestriction": {"base": {"value": 0}, "final": {"value": 0}},
                "naturalPenalty": {"base": {"value": 0}, "final": {"value": 0}},
                "wearArmorRequirement": {"base": {"value": 0}, "final": {"value": 0}},
                "isEnchanted": {"value": False},
                "type": {"value": "natural"},
                "localization": {"value": "complete"},
                "quality": {"value": 0},
                "equipped": {"value": True},
            },
        }
        items.append(armor)

    # Notes
    notes = []
    note_fields = [
        ("advantages", "Ventajas y desventajas"),
        ("natural_abilities", "Habilidades naturales"),
        ("essential_abilities", "Habilidades esenciales"),
        ("powers", "Poderes"),
        ("special", "Especial"),
        ("ki_skills", "Habilidades de Ki"),
        ("techniques", "Técnicas"),
        ("magic_level", "Nivel de Magia"),
        ("psychic_disciplines", "Disciplinas Psíquicas"),
        ("psychic_powers", "Poderes Psíquicos"),
    ]
    for idx, (field, label) in enumerate(note_fields, 1):
        val = parsed.get(field, "")
        if val:
            notes.append({
                "_id": stable_id(actor_id, "note", str(idx)),
                "type": "note",
                "name": f"{label}: {val}",
                "system": {},
            })
    system["general"]["notes"] = notes

    sight_range = parsed["perception"] * 20

    return {
        "_id": actor_id,
        "name": parsed["name"],
        "type": "character",
        "img": "icons/svg/mystery-man.svg",
        "prototypeToken": {
            "name": parsed["name"],
            "displayName": 0,
            "actorLink": False,
            "width": 1, "height": 1,
            "lockRotation": False, "rotation": 0, "alpha": 1,
            "disposition": 1, "displayBars": 0,
            "bar1": {"attribute": "characteristics.secondaries.lifePoints"},
            "bar2": {"attribute": None},
            "flags": {"levels": {"tokenHeight": 0}},
            "randomImg": False,
            "light": {"dim": 0, "bright": 0, "angle": 360, "color": None, "alpha": 0.25,
                      "animation": {"speed": 5, "intensity": 5, "type": None, "reverse": False},
                      "coloration": 1, "attenuation": 0.5, "luminosity": 0.5, "saturation": 0,
                      "contrast": 0, "shadows": 0, "darkness": {"min": 0, "max": 1}},
            "texture": {"src": "icons/svg/mystery-man.svg", "tint": None,
                        "scaleX": 1, "scaleY": 1, "offsetX": 0, "offsetY": 0, "rotation": 0},
            "sight": {"angle": 360, "enabled": sight_range > 0, "range": sight_range,
                      "brightness": 1, "visionMode": "basic", "attenuation": 0.1,
                      "saturation": 0, "contrast": 0},
            "appendNumber": False, "prependAdjective": False, "detectionModes": [],
        },
        "items": items,
        "effects": [],
        "folder": None,
        "sort": 0,
        "flags": {"core": {}, "animu-exxet": {"sourceFile": parsed["source_file"]}},
        "system": system,
        "_stats": {"systemId": "animabf", "systemVersion": "2.2.1", "coreVersion": "14.359",
                   "createdTime": 1713571200000, "modifiedTime": 1713571200000, "lastModifiedBy": "animuExxetGen001"},
        "ownership": {"default": 0},
        "_key": f"!actors!{actor_id}",
    }


def main():
    parser = argparse.ArgumentParser(description="Convierte fichas Excel de Anima a JSON de animabf.")
    parser.add_argument("files", nargs="+", help="Archivos .xlsm a convertir")
    parser.add_argument("-o", "--output", default=".", help="Directorio de salida")
    args = parser.parse_args()

    template = json.loads(TEMPLATE_PATH.read_text(encoding="utf-8"))
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    for filepath in args.files:
        path = Path(filepath)
        if not path.exists():
            print(f"No existe: {path}", file=sys.stderr)
            continue
        try:
            parsed = parse_excel(path)
            actor = build_actor(parsed, template)
            out_path = output_dir / f"{path.stem}.json"
            out_path.write_text(json.dumps(actor, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
            print(f"{path.name} -> {out_path.name}: {parsed['name']} (Nivel {parsed['level']})")
        except Exception as e:
            print(f"Error en {path.name}: {e}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
